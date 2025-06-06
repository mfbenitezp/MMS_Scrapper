import os
import time
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


def setup_driver():
    """Setup Edge driver with appropriate options"""
    edge_options = Options()
    edge_options.add_argument("--no-sandbox")
    edge_options.add_argument("--disable-dev-shm-usage")
    edge_options.add_argument("--window-size=1920,1080")
    service = Service(EdgeChromiumDriverManager().install())
    return webdriver.Edge(service=service, options=edge_options)


def manual_login(driver, test_url):
    """Handle manual authentication process"""
    print("=== Manual Authentication ===")
    print("1. A browser window will open")
    print("2. Please log in to St Andrews manually")
    print("3. Navigate to the grades page to confirm you're logged in")
    print("4. Come back here and press Enter when authentication is complete")
    print()
    
    driver.get(test_url)
    input("➡️ Once you're logged in and see the grades table, press Enter here...")
    
    # Verify authentication
    try:
        if "login" in driver.current_url.lower() or "auth" in driver.current_url.lower():
            print("Warning: Still appears to be on login page")
            input("Please complete login and press Enter again...")
        print("Authentication verified! Starting data extraction...")
        return True
    except Exception as e:
        print(f"Note: {e}")
        print("Proceeding anyway...")
        return True


def extract_table_html(driver):
    """Extract the grades table HTML"""
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "gradesTable"))
    )
    table_element = driver.find_element(By.ID, "gradesTable")
    return table_element.get_attribute("outerHTML")


def parse_html_table_to_dataframe(table_html):
    """Parse HTML table into pandas DataFrame"""
    soup = BeautifulSoup(table_html, "html.parser")
    table = soup.find("table")
    df = pd.read_html(str(table), header=[0, 1])[0]  # Read MultiIndex headers
    return df


def filter_grades_dataframe(df, module_code):
    """Filter and process grades DataFrame"""
    print(f"📋 Processing columns for {module_code}...")

    try:
        filtered_df = df[[('Student ↓↑', 'Matric Number ↓↑'), ('Result ↓↑', 'Calc Grade ↓↑')]]
    except KeyError as e:
        print(f"❌ Columns not found in {module_code}. Error: {e}")
        return None, None

    filtered_df.columns = ['Matric Number', 'Calc Grade']

    # Split data from summary rows
    student_data = filtered_df.iloc[:-6].copy()
    summary_data = filtered_df.iloc[-6:].copy()

    # Convert grades to numeric (ignore non-numeric or missing values)
    student_data['Calc Grade'] = pd.to_numeric(student_data['Calc Grade'], errors='coerce')

    # Calculate extra statistics
    total_students = student_data['Calc Grade'].count()
    pct_gte_16_5 = (student_data['Calc Grade'] >= 16.5).sum() / total_students * 100 if total_students > 0 else 0
    pct_14_16 = ((student_data['Calc Grade'] >= 14) & (student_data['Calc Grade'] < 16)).sum() / total_students * 100 if total_students > 0 else 0

    # Format summary row from table
    summary_row = summary_data.set_index('Matric Number').T
    summary_row.columns.name = None
    summary_row['Module'] = module_code

    # Add new calculated percentages
    summary_row['% ≥ 16.5'] = round(pct_gte_16_5, 2)
    summary_row['% between 14–16'] = round(pct_14_16, 2)

    summary_row = summary_row.set_index('Module')

    return student_data, summary_row


def save_charts_as_png(driver, module_code, charts_dir="charts"):
    """Save both scatter charts for a module"""
    saved_count = 0
    
    # URLs for the two different scatter charts
    urls = {
        "GraphPage": f"https://mms.st-andrews.ac.uk/mms/module/2024_5/S2/{module_code}/Final+grade/GraphPage",
        "SubmitResults": f"https://mms.st-andrews.ac.uk/mms/module/2024_5/S2/{module_code}/Final+grade/SubmitResults"
    }
    
    # Create folder for this module
    folder_path = os.path.join(charts_dir, module_code)
    os.makedirs(folder_path, exist_ok=True)
    
    for chart_type, url in urls.items():
        try:
            print(f"  Loading {chart_type} chart for {module_code}...")
            driver.get(url)
            time.sleep(3)  # Wait for page to load
            
            # Check if we got redirected to login
            if "login" in driver.current_url.lower() or "auth" in driver.current_url.lower():
                print(f"  Session expired! Please re-authenticate.")
                input("Complete authentication and press Enter...")
                driver.get(url)
                time.sleep(5)
            
            # Look for scatter chart
            try:
                wait = WebDriverWait(driver, 15)
                
                if chart_type == "GraphPage":
                    # Scatter chart 1 from GraphPage
                    scatter_chart = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#scatterChart .user-select-none.svg-container")))
                    chart_name = "ScatterChart_1"
                else:
                    # Scatter chart 2 from SubmitResults
                    scatter_chart = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#scatterChart .user-select-none.svg-container")))
                    chart_name = "ScatterChart_2"
                
                print(f"    Found {chart_name} for {module_code}")
                
                # Give chart extra time to fully render
                time.sleep(5)
                
                # Scroll to chart to ensure it's visible
                driver.execute_script("arguments[0].scrollIntoView(true);", scatter_chart)
                time.sleep(2)
                
                # Take screenshot of the chart
                filename = os.path.join(folder_path, f"{chart_name}.png")
                scatter_chart.screenshot(filename)
                
                print(f"    ✓ Saved {chart_name}.png")
                saved_count += 1
                
            except Exception as e:
                print(f"    ✗ {chart_name} not found for {module_code}: {e}")
        
        except Exception as e:
            print(f"  Error processing {chart_type} for {module_code}: {e}")
    
    return saved_count


def add_charts_to_excel(wb, module_code, charts_dir="charts"):
    """Add charts to the existing module sheet in the workbook"""
    try:
        # Get the existing sheet for this module
        if module_code in wb.sheetnames:
            sheet = wb[module_code]
        else:
            return False
        
        # Find the last row with data to position charts below
        last_row = sheet.min_row   # Add some spacing
        
        # Look for chart files for this module
        module_path = os.path.join(charts_dir, module_code)
        if not os.path.exists(module_path):
            return False
        
        chart_files = [f for f in os.listdir(module_path) if f.endswith('.png') and not f.endswith('_resized.png')]
        
        if not chart_files:
            return False
        
        # Sort chart files for consistent ordering
        chart_files.sort()
        
        current_row = last_row
        for chart_file in chart_files:
            chart_path = os.path.join(module_path, chart_file)
            
            try:
                # Check if original file exists
                if not os.path.exists(chart_path):
                    print(f"    ✗ Chart file not found: {chart_path}")
                    continue
                
                # Resize image to reasonable size for Excel
                img = PILImage.open(chart_path)
                img.thumbnail((800, 600))  # Resize to max dimensions
                temp_path = chart_path.replace(".png", "_resized.png")
                img.save(temp_path)
                
                # Verify temp file was created
                if not os.path.exists(temp_path):
                    print(f"    ✗ Failed to create resized image: {temp_path}")
                    continue
                
                # Add to Excel sheet
                xl_img = XLImage(temp_path)
                cell_location = f"D{current_row}"
                sheet.add_image(xl_img, cell_location)
                current_row += 25  # Space between charts
                
                print(f"    ✓ Added {chart_file} to {module_code} sheet")
   
            except Exception as e:
                print(f"    ✗ Failed to add {chart_file} to {module_code} sheet: {e}")
        
        return True
        
    except Exception as e:
        print(f"  Error adding charts to {module_code} sheet: {e}")
        return False


def main():
    """Main function"""
    print("St Andrews Module Data and Charts Extractor")
    print("=" * 60)
    
    # Module codes to process
    module_codes = ['GG4258', 'GG3281', 'GG1002', 'GG2014', 'GG4248', 'GG4247', 'SS5103',
                    'GG4254', 'GG4257', 'GG3205', 'GG3213', 'GG3214', 'GG5005', 'GG4399',
                    'SD4126', 'SD4129', 'SD4133', 'SD1004', 'SD4225', 'SD2006', 'SD2100',
                    'SD4110', 'SD3102', 'SD3101', 'SD4120', 'SD4125', 'SD4297', 'SD5801',
                    'SD5802', 'SD5805', 'SD5806', 'SD5807', 'SD5810', 'SD5820', 'SD5821',
                    'SD5811', 'SD5813', 'SD5812']
    
    base_url = "https://mms.st-andrews.ac.uk/mms/module/2024_5/S2"
    output_filename = "Complete_Modules_Data_and_Charts.xlsx"
    charts_dir = "charts"
    
    # Create charts directory
    os.makedirs(charts_dir, exist_ok=True)
    
    # Setup driver
    driver = setup_driver()
    all_grades = {}
    all_summaries = []
    
    try:
        # Step 1: Manual login
        login_url = f"{base_url}/{module_codes[0]}/Final+grade/"
        manual_login(driver, login_url)
        
        print(f"\n🔍 Processing {len(module_codes)} modules...")
        print("=" * 40)
        
        # Step 2: Extract grades data and charts for each module
        total_charts_saved = 0
        successful_modules = 0
        
        for i, module_code in enumerate(module_codes, 1):
            print(f"\n[{i}/{len(module_codes)}] Processing module {module_code}...")
            
            # Extract grades data
            try:
                module_url = f"{base_url}/{module_code}/Final+grade/"
                driver.get(module_url)
                
                table_html = extract_table_html(driver)
                df = parse_html_table_to_dataframe(table_html)
                student_data, summary_row = filter_grades_dataframe(df, module_code)
                
                if student_data is not None:
                    all_grades[module_code] = student_data
                    all_summaries.append(summary_row)
                    print(f"  ✅ Grades data collected for {module_code}")
                else:
                    print(f"  ⚠️ No grades data for {module_code}")
                    
            except Exception as e:
                print(f"  ⚠️ Error extracting grades for {module_code}: {e}")
            
            # Extract charts
            try:
                charts_saved = save_charts_as_png(driver, module_code, charts_dir)
                if charts_saved > 0:
                    total_charts_saved += charts_saved
                    successful_modules += 1
                    print(f"  ✅ {charts_saved} charts saved for {module_code}")
                else:
                    print(f"  ⚠️ No charts saved for {module_code}")
            except Exception as e:
                print(f"  ⚠️ Error extracting charts for {module_code}: {e}")
            
            # Small delay between modules
            time.sleep(2)
        
        # Step 3: Create Excel workbook with grades data
        if all_grades:
            print(f"\n📊 Creating Excel workbook with grades and charts...")
            
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                # Write individual module sheets with grades
                for module_code, df in all_grades.items():
                    df.to_excel(writer, sheet_name=module_code, index=False)
                
                # Write summary sheet
                if all_summaries:
                    summary_df = pd.concat(all_summaries)
                    summary_df.to_excel(writer, sheet_name="Summary")
            
            # Step 4: Add charts to the existing workbook
            print("📈 Adding charts to Excel sheets...")
            wb = load_workbook(output_filename)
            
            charts_added = 0
            for module_code in all_grades.keys():
                if add_charts_to_excel(wb, module_code, charts_dir):
                    charts_added += 1
            
            wb.save(output_filename)
            
            # Final summary
            print("\n" + "=" * 60)
            print("EXTRACTION COMPLETE!")
            print(f"Processed modules: {len(module_codes)}")
            print(f"Modules with grades: {len(all_grades)}")
            print(f"Modules with charts: {successful_modules}")
            print(f"Total charts saved: {total_charts_saved}")
            print(f"Sheets with charts added: {charts_added}")
            print(f"Output file: {output_filename}")
            print("=" * 60)
            
        else:
            print("❌ No grades data collected. Please check authentication and module URLs.")
    
    except KeyboardInterrupt:
        print("\nProcess interrupted by user")
    except Exception as e:
        print(f"Unexpected error: {e}")
        #import traceback
        #print("Full error traceback:")
        #traceback.print_exc()
    finally:
        input("\nPress Enter to close the browser...")
        driver.quit()
        
        # Clean up any remaining temporary chart files
        print("Cleaning up temporary files...")
        try:
            for module_code in module_codes:
                module_path = os.path.join(charts_dir, module_code)
                if os.path.exists(module_path):
                    for file in os.listdir(module_path):
                        if file.endswith('_resized.png'):
                            temp_file_path = os.path.join(module_path, file)
                            try:
                                if os.path.exists(temp_file_path):
                                    os.remove(temp_file_path)
                                    print(f"  Removed: {temp_file_path}")
                            except Exception as cleanup_error:
                                print(f"  Could not remove {temp_file_path}: {cleanup_error}")
        except Exception as e:
            print(f"Note: Error during cleanup: {e}")

if __name__ == "__main__":
    main()